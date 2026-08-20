"""
build_dataset.py
================
THE PIPELINE. Runs once. Rebuilds the dataset from the raw ONS spreadsheet.

    ONS Excel  ->  two tidy tables  ->  joined  ->  data/cost_of_living_merged.csv

Run it from the terminal:   python build_dataset.py
Then analysis.ipynb and app.py both just READ the merged CSV.
They never re-do this work.

Requires:  pip install openpyxl pandas duckdb
"""

from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
import duckdb

# ---- CHANGE THIS to wherever the ONS file lives on your laptop -------------
SRC = Path('consumerpriceinflationdetailedreferencetables.xlsx')
OUT = Path('data')
OUT.mkdir(exist_ok=True)

wb = load_workbook(SRC, read_only=True, data_only=True)


# ===========================================================================
# 1. FACT TABLE  <-  ONS Table 8 (CPIH: 12-month % change by division)
# ===========================================================================
ws8 = wb['Table 8']
rows8 = list(ws8.iter_rows(min_row=1, max_row=21, values_only=True))

years  = rows8[5]      # the row holding 2025 / 2026
months = rows8[6]      # the row holding Jan / Feb / Mar ...

MONTH_NAMES = ('Jan','Feb','Mar','Apr','May','Jun',
               'Jul','Aug','Sep','Oct','Nov','Dec')

# Find every column that is a real month, then keep the last three.
month_cols = [(i, f"{years[i]}-{months[i]}")
              for i in range(len(months)) if months[i] in MONTH_NAMES]
last3 = month_cols[-3:]

fact = []
for r in rows8[8:20]:                       # the 12 divisions
    code = str(r[2]).strip().split()[0]     # "01", "02", ...
    for col_index, month_label in last3:
        fact.append({'division_code': code,
                     'month':         month_label,
                     'annual_rate':   r[col_index]})
fact = pd.DataFrame(fact)


# ===========================================================================
# 2. DIMENSION TABLE  <-  ONS Table 9 (CPIH weights)
# ===========================================================================

ws9 = wb['Table 9']
rows9 = list(ws9.iter_rows(min_row=1, max_row=21, values_only=True))

newest_weights_col = len(rows9[6]) - 1      # the last column = most recent year

dim = []                                     # <- a LIST (loose rows)
for r in rows9[8:20]:
    code, name = str(r[2]).strip().split(None, 1)
    raw = r[newest_weights_col]

    if not isinstance(raw, (int, float)):
        raise ValueError(f"Bad weight for division {code}: {raw!r} — check Table 9")

    dim.append({'division_code': code,
                'division_name': name.strip(),
                'weight':        round(float(raw), 1)})

dim = pd.DataFrame(dim)                      # <- NOW a DATAFRAME. Don't lose this line.)

# ===========================================================================
# 3. JOIN THEM  ->  the merged dataset
# ===========================================================================
fact.to_csv(OUT / 'inflation_by_category.csv', index=False)
dim.to_csv(OUT / 'category_weights.csv',       index=False)

con = duckdb.connect()
con.execute(f"CREATE TABLE fact AS SELECT * FROM read_csv_auto('{OUT}/inflation_by_category.csv')")
con.execute(f"CREATE TABLE dim  AS SELECT * FROM read_csv_auto('{OUT}/category_weights.csv')")

con.execute("""
CREATE TABLE merged AS
SELECT
    f.month,
    f.division_code,
    d.division_name,
    f.annual_rate,
    d.weight,

    -- the whole point of the join: rate x weight = real impact on a budget
    ROUND((d.weight / 1000.0) * f.annual_rate, 2)          AS weighted_contribution,

    -- month-on-month change, via a SQL window function
    -- NOTE: this is currently WRONG. See Q5 in the Analysis Guide.
    ROUND(f.annual_rate - LAG(f.annual_rate) OVER (
          PARTITION BY f.division_code ORDER BY f.month), 1) AS mom_change

FROM fact f
JOIN dim  d ON f.division_code = d.division_code
""")

con.execute(f"COPY merged TO '{OUT}/cost_of_living_merged.csv' (HEADER, DELIMITER ',')")


# ===========================================================================
# 4. Confirm it worked (a pipeline should tell you it succeeded)
# ===========================================================================
merged = con.execute("SELECT * FROM merged").df()

print(f"Months extracted : {[m[1] for m in last3]}")
print(f"Fact rows        : {len(fact)}   (12 categories x 3 months)")
print(f"Dimension rows   : {len(dim)}")
print(f"Weights sum to   : {dim.weight.sum():.1f}   (should be ~1000)")
print(f"Merged rows      : {len(merged)}")
print(f"Written to       : {OUT}/cost_of_living_merged.csv")
